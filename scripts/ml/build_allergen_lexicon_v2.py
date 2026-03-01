#!/usr/bin/env python3
"""Build a comprehensive allergen lexicon (Big 9 + cereals-with-gluten)."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, MutableMapping, Sequence, Set, Tuple

import openpyxl


WORKBOOK_SHEET_MAP = "Allergen-Ingredient Map"
WORKBOOK_SHEET_ALIAS_POLICY = "Lexicon Alias Policy"
POLICY_CLASS_ANY = "*"
WORKBOOK_ALLERGEN_TO_CLASS: Mapping[str, str] = {
    "Milk": "milk",
    "Egg": "egg",
    "Peanut": "peanut",
    "Tree Nut": "tree_nut",
    "Wheat": "wheat",
    "Soy": "soy",
    "Fish": "fish",
    "Crustacean Shellfish": "shellfish",
    "Sesame": "sesame",
}

DEFAULT_DATASET_FILES: Sequence[str] = (
    "ml/data/processed/usda_only_train.jsonl",
    "ml/data/processed/usda_only_val.jsonl",
    "ml/data/processed/usda_only_holdout.jsonl",
    "ml/data/processed/openfoodfacts_targeted_examples.jsonl",
)

CLASS_ORDER: Sequence[str] = (
    "milk",
    "egg",
    "peanut",
    "tree_nut",
    "wheat",
    "soy",
    "fish",
    "shellfish",
    "sesame",
    "cereals_with_gluten",
)

CLASS_DISPLAY: Mapping[str, str] = {
    "milk": "Milk",
    "egg": "Egg",
    "peanut": "Peanut",
    "tree_nut": "Tree Nut",
    "wheat": "Wheat",
    "soy": "Soy",
    "fish": "Fish",
    "shellfish": "Crustacean Shellfish",
    "sesame": "Sesame",
    "cereals_with_gluten": "Cereals with Gluten",
}

CLASS_SCOPE: Mapping[str, str] = {
    "milk": "big9",
    "egg": "big9",
    "peanut": "big9",
    "tree_nut": "big9",
    "wheat": "big9",
    "soy": "big9",
    "fish": "big9",
    "shellfish": "big9",
    "sesame": "big9",
    "cereals_with_gluten": "cereals_with_gluten",
}

CLASS_LABEL_KEYS: Mapping[str, Set[str]] = {
    "milk": {"milk"},
    "egg": {"egg"},
    "peanut": {"peanut"},
    "tree_nut": {"tree nut"},
    "wheat": {"wheat"},
    "soy": {"soy"},
    "fish": {"fish"},
    "shellfish": {"shellfish"},
    "sesame": {"sesame"},
    # USDA/OpenFoodFacts labels in this corpus do not include a dedicated
    # "cereals_with_gluten" class. Use wheat labels only as proxy.
    "cereals_with_gluten": {"wheat"},
}

CLASS_ROOT_TOKENS: Mapping[str, Set[str]] = {
    "milk": {"milk", "whey", "casein", "caseinate", "cream", "butter", "cheese", "curd", "lact", "rennet", "yogurt"},
    "egg": {"egg", "albumin", "albumen", "ovalbumin", "ovomucoid", "ovotransferrin", "ovomucin", "lysozyme", "meringue", "mayonnaise"},
    "peanut": {"peanut", "peanuts", "groundnut", "groundnuts", "goober", "arachis"},
    "tree_nut": {
        "almond",
        "cashew",
        "pecan",
        "walnut",
        "pistachio",
        "hazelnut",
        "filbert",
        "macadamia",
        "brazil",
        "chestnut",
        "pine",
        "pignoli",
        "pinyon",
        "nutella",
        "marzipan",
        "gianduja",
        "coconut",
    },
    "wheat": {"wheat", "flour", "semolina", "durum", "bulgur", "couscous", "spelt", "triticale", "seitan", "kamut", "farro", "einkorn", "emmer", "gluten"},
    "soy": {"soy", "soya", "soybean", "soybeans", "tofu", "tempeh", "miso", "natto", "edamame", "tamari", "shoyu", "lecithin"},
    "fish": {"fish", "anchovy", "anchovies", "salmon", "tuna", "pollock", "cod", "trout", "herring", "mackerel", "sardine", "sardines"},
    "shellfish": {"shellfish", "shrimp", "prawn", "crab", "lobster", "krill", "scampi", "crawfish", "crayfish", "clam", "clams", "mussel", "oyster", "scallop", "surimi"},
    "sesame": {"sesame", "tahini", "tahina", "tehina", "gingelly", "benne", "benniseed", "sesamum", "gomasio", "halvah", "halva"},
    "cereals_with_gluten": {
        "wheat",
        "barley",
        "rye",
        "oat",
        "oats",
        "triticale",
        "spelt",
        "kamut",
        "farro",
        "einkorn",
        "emmer",
        "durum",
        "semolina",
        "malt",
        "gluten",
        "seitan",
        "matzo",
        "matzah",
        "matza",
    },
}

MANUAL_CANONICAL_SEEDS: Mapping[str, Mapping[str, Sequence[str]]] = {
    "milk": {
        "Whey Protein Hydrolysate": ("Whey Protein Hydrolysate", "Hydrolyzed Whey Protein", "WPH"),
        "Milk Solids": ("Milk Solids", "Milk Solids Nonfat", "Nonfat Dry Milk", "Skim Milk Powder"),
        "Caseinate Salts": ("Sodium Caseinate", "Calcium Caseinate", "Potassium Caseinate"),
    },
    "egg": {
        "Egg Powder": ("Egg Powder", "Whole Egg Powder", "Dried Whole Egg"),
        "Liquid Egg": ("Liquid Egg", "Liquid Whole Egg", "Pasteurized Egg"),
    },
    "tree_nut": {
        "Almond Milk": ("Almond Milk", "Almondmilk"),
        "Coconut Milk": ("Coconut Milk", "Coconutmilk"),
    },
    "soy": {
        "Soybean Lecithin": ("Soybean Lecithin", "Soy Lecithin"),
        "Defatted Soy Flour": ("Defatted Soy Flour", "Defatted Soybean Flour"),
        "Soy Isolate": ("Soy Isolate", "Soy Protein Isolate", "Soybean Protein Isolate"),
    },
    "fish": {
        "Sardine": ("Sardine", "Sardines"),
        "Mackerel": ("Mackerel",),
    },
    "shellfish": {
        "Crab Meat": ("Crab Meat", "Crabmeat"),
        "Lobster Extract": ("Lobster Extract",),
    },
    "cereals_with_gluten": {
        "Barley": ("Barley", "Barley Flour", "Pearled Barley", "Barley Malt", "Malted Barley", "Barley Extract"),
        "Rye": ("Rye", "Rye Flour", "Rye Meal", "Rye Malt"),
        "Oat": ("Oat", "Oats", "Oat Flour", "Rolled Oats", "Oat Bran", "Oatmeal"),
        "Spelt": ("Spelt", "Spelt Flour"),
        "Kamut": ("Kamut", "Khorasan Wheat", "Kamut Flour"),
        "Farro": ("Farro",),
        "Einkorn": ("Einkorn", "Einkorn Flour"),
        "Emmer": ("Emmer",),
        "Durum": ("Durum", "Durum Flour", "Durum Wheat", "Durum Semolina"),
        "Semolina": ("Semolina", "Semolina Flour"),
        "Triticale": ("Triticale",),
        "Gluten": ("Gluten", "Vital Wheat Gluten", "Wheat Gluten", "Seitan"),
        "Malt": ("Malt", "Malted", "Malt Extract", "Malt Syrup", "Malted Barley Extract"),
    },
}


TOKEN_RE = re.compile(r"[a-z0-9]+")
DISPLAY_SPACE_RE = re.compile(r"\s+")
PAREN_RE = re.compile(r"\(([^()]*)\)")
PAREN_SPLIT_RE = re.compile(r"[,/;]")
PAREN_PREFIX_RE = re.compile(
    r"^(all forms:?|all types:?|incl\.?|including|includes?|e\.g\.?|if [a-z -]+derived|and meal)\s+",
    flags=re.IGNORECASE,
)

CANDIDATE_SPLIT_RE = re.compile(r"[;\(\)\[\]\{\}]")
CANDIDATE_NOISE_PHRASES = {
    "contains",
    "ingredients",
    "ingredient",
    "or less",
    "and or",
    "percent or less",
    "this product",
    "may contain",
}

STOPWORDS = {
    "a",
    "an",
    "and",
    "or",
    "of",
    "to",
    "for",
    "with",
    "from",
    "in",
    "on",
    "at",
    "contains",
    "ingredients",
    "ingredient",
    "less",
    "than",
    "percent",
    "natural",
    "flavor",
    "flavors",
    "spice",
    "spices",
    "artificial",
}

GENERIC_RULE_VARIANT_BLOCKLIST = {
    "oil",
    "oils",
    "protein",
    "proteins",
    "extract",
    "extracts",
    "flour",
    "powder",
    "meal",
    "seed",
    "seeds",
    "milk",
    "cream",
    "fish",
    "shellfish",
    "egg",
    "soy",
    "wheat",
    "sesame",
    "nut",
    "nuts",
}

DEFAULT_AMBIGUOUS_ALIAS_TOKENS: Set[Tuple[str, ...]] = {
    ("extract",),
    ("extracts",),
    ("flavor",),
    ("flavoring",),
    ("flavorings",),
    ("flavors",),
    ("spice",),
    ("spices",),
    ("seasoning",),
    ("seasonings",),
    ("ingredient",),
    ("ingredients",),
    ("natural", "flavor"),
    ("natural", "flavors"),
    ("artificial", "flavor"),
    ("artificial", "flavors"),
}

VALID_ALIAS_POLICY_ACTIONS = {"allow", "deny", "review"}

TRIE_TERM_IDS_KEY = "__alias_ids__"


@dataclass(frozen=True)
class CanonicalEntry:
    canonical_id: str
    class_key: str
    canonical_name: str
    source_type: str
    source_origin: str


@dataclass(frozen=True)
class AliasEntry:
    alias_id: int
    canonical_id: str
    class_key: str
    alias_display: str
    alias_tokens: Tuple[str, ...]
    source_type: str
    source_origin: str
    status: str


@dataclass
class AliasStats:
    rows_matched: int = 0
    target_rows: int = 0
    other_labeled_rows: int = 0
    mentions_total: int = 0


@dataclass
class CanonicalStats:
    rows_matched: int = 0
    target_rows: int = 0
    mentions_total: int = 0


@dataclass
class ClassSummary:
    labeled_rows: int = 0
    labeled_rows_with_alias: int = 0
    proxy_rows: int = 0
    proxy_rows_with_alias: int = 0
    rows_with_alias_any: int = 0
    alias_mentions_total: int = 0
    unique_alias_ids: Set[int] | None = None

    def __post_init__(self) -> None:
        if self.unique_alias_ids is None:
            self.unique_alias_ids = set()


@dataclass(frozen=True)
class CandidateRow:
    class_key: str
    candidate_display: str
    candidate_tokens: Tuple[str, ...]
    support_rows: int
    target_rows: int
    unmatched_target_rows: int
    other_labeled_rows: int
    unlabeled_rows: int
    precision: float
    exclusivity: float
    score: float
    recommendation: str
    reason: str


@dataclass(frozen=True)
class AliasPolicyRow:
    class_key: str
    alias_tokens: Tuple[str, ...]
    action: str
    reason: str
    source_origin: str


@dataclass(frozen=True)
class AliasActionRow:
    alias_id: int
    class_key: str
    canonical_id: str
    alias_display: str
    alias_tokens: Tuple[str, ...]
    source_type: str
    source_origin: str
    status_before: str
    status_after: str
    action: str
    reason: str
    rows_matched: int
    target_rows: int
    other_labeled_rows: int
    precision: float
    exclusivity: float


@dataclass(frozen=True)
class CoverageGapRow:
    class_key: str
    scope: str
    target_rows: int
    covered_rows: int
    uncovered_rows: int
    coverage_pct: float


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build Big 9 + cereals-with-gluten allergen lexicon.")
    parser.add_argument(
        "--workbook-input",
        default="/Users/mattdavis/Desktop/allergen_ingredient_database.xlsx",
        help="Input workbook path.",
    )
    parser.add_argument(
        "--workbook-output",
        default="/Users/mattdavis/Desktop/allergen_ingredient_database_lexicon_v2.xlsx",
        help="Output workbook path.",
    )
    parser.add_argument(
        "--dataset",
        action="append",
        default=[],
        help="JSONL dataset path (repeatable). Defaults to processed USDA+OFF files.",
    )
    parser.add_argument(
        "--output-dir",
        default="ml/data/analysis/lexicon_v2",
        help="Directory for CSV outputs.",
    )
    parser.add_argument(
        "--denylist-csv",
        default="",
        help="Optional CSV with alias policy rows: class,alias,action,reason.",
    )
    parser.add_argument(
        "--min-candidate-support",
        type=int,
        default=20,
        help="Minimum support rows for candidate review queue.",
    )
    parser.add_argument(
        "--min-candidate-precision",
        type=float,
        default=0.75,
        help="Minimum target precision for Big 9 candidates.",
    )
    parser.add_argument(
        "--max-review-per-class",
        type=int,
        default=150,
        help="Max candidate rows per class in review queue.",
    )
    parser.add_argument(
        "--candidate-min-unmatched-target-rows",
        type=int,
        default=8,
        help="Minimum unmatched positive rows required for candidate review rows.",
    )
    parser.add_argument(
        "--candidate-min-exclusivity",
        type=float,
        default=0.55,
        help="Minimum exclusivity (target / (target + other-labeled)) for Big 9 review candidates.",
    )
    parser.add_argument(
        "--auto-demote-min-support",
        type=int,
        default=400,
        help="Auto-demote active aliases with at least this many matched rows when precision/exclusivity are weak.",
    )
    parser.add_argument(
        "--auto-demote-max-precision",
        type=float,
        default=0.30,
        help="Auto-demote aliases at or below this precision threshold.",
    )
    parser.add_argument(
        "--auto-demote-max-exclusivity",
        type=float,
        default=0.45,
        help="Auto-demote aliases at or below this exclusivity threshold.",
    )
    return parser.parse_args()


def normalize_ascii(value: str) -> str:
    safe = unicodedata.normalize("NFKD", value or "")
    return safe.encode("ascii", "ignore").decode("ascii")


def normalize_display(value: str) -> str:
    safe = normalize_ascii(value)
    safe = safe.replace("®", "").replace("™", "")
    safe = safe.replace("’", "'")
    safe = DISPLAY_SPACE_RE.sub(" ", safe).strip()
    safe = safe.strip(",;:.")
    return safe


def tokenize(value: str) -> Tuple[str, ...]:
    safe = normalize_ascii(value).lower()
    return tuple(TOKEN_RE.findall(safe))


def title_from_tokens(tokens: Sequence[str]) -> str:
    return " ".join(tokens)


def slugify(value: str) -> str:
    safe = normalize_ascii(value).lower()
    safe = re.sub(r"[^a-z0-9]+", "_", safe).strip("_")
    safe = re.sub(r"_+", "_", safe)
    return safe or "term"


def allergen_key_from_label(value: str) -> str:
    safe = normalize_display(value).lower().replace("-", " ")
    safe = re.sub(r"\s+", " ", safe)
    mapping = {
        "milk": "milk",
        "egg": "egg",
        "peanut": "peanut",
        "tree nut": "tree_nut",
        "tree nuts": "tree_nut",
        "wheat": "wheat",
        "soy": "soy",
        "fish": "fish",
        "shellfish": "shellfish",
        "crustacean shellfish": "shellfish",
        "sesame": "sesame",
    }
    return mapping.get(safe, "")


def class_key_from_policy_label(value: str) -> str:
    safe = normalize_display(value).lower().replace("-", " ").replace("_", " ")
    safe = re.sub(r"\s+", " ", safe).strip()
    if safe in {"*", "all", "any", "global"}:
        return POLICY_CLASS_ANY

    mapping = {
        "milk": "milk",
        "egg": "egg",
        "peanut": "peanut",
        "tree nut": "tree_nut",
        "tree nuts": "tree_nut",
        "tree_nut": "tree_nut",
        "wheat": "wheat",
        "soy": "soy",
        "fish": "fish",
        "shellfish": "shellfish",
        "crustacean shellfish": "shellfish",
        "sesame": "sesame",
        "cereals with gluten": "cereals_with_gluten",
        "cereals_with_gluten": "cereals_with_gluten",
    }
    return mapping.get(safe, "")


def alias_precision(stat: AliasStats) -> float:
    return (float(stat.target_rows) / float(stat.rows_matched)) if stat.rows_matched else 0.0


def alias_exclusivity(stat: AliasStats) -> float:
    labeled_rows = stat.target_rows + stat.other_labeled_rows
    return (float(stat.target_rows) / float(labeled_rows)) if labeled_rows else 0.0


def class_target_hit(class_key: str, row_labels: Set[str], wheat_proxy: bool) -> bool:
    if class_key == "cereals_with_gluten":
        return wheat_proxy
    return class_key in row_labels


def has_other_non_target_label(class_key: str, row_labels: Set[str]) -> bool:
    if not row_labels:
        return False
    if class_key == "cereals_with_gluten":
        return any(label != "wheat" for label in row_labels)
    return any(label != class_key for label in row_labels)


def resolve_policy_for_alias(
    class_key: str,
    alias_tokens: Tuple[str, ...],
    policy_rows: Mapping[Tuple[str, Tuple[str, ...]], AliasPolicyRow],
) -> AliasPolicyRow | None:
    specific = policy_rows.get((class_key, alias_tokens))
    if specific:
        return specific
    return policy_rows.get((POLICY_CLASS_ANY, alias_tokens))


def split_slash_aliases(value: str) -> List[str]:
    return [item.strip() for item in re.split(r"\s*/\s*", value or "") if item.strip()]


def base_alias_without_parentheses(alias_text: str) -> str:
    return normalize_display(PAREN_RE.sub("", alias_text))


def clean_parenthetical_fragment(fragment: str) -> str:
    safe = normalize_display(fragment)
    safe = PAREN_PREFIX_RE.sub("", safe)
    safe = safe.replace("etc.", "").replace("etc", "")
    safe = safe.strip(" ,;:.")
    return safe


def parenthetical_alias_expansions(alias_text: str) -> Set[str]:
    out: Set[str] = set()
    base = base_alias_without_parentheses(alias_text)
    base_tokens = tokenize(base)
    if not base_tokens:
        return out

    for match in PAREN_RE.finditer(alias_text):
        inner = match.group(1)
        for part in PAREN_SPLIT_RE.split(inner):
            piece = clean_parenthetical_fragment(part)
            if not piece:
                continue
            piece_tokens = tokenize(piece)
            if not piece_tokens or len(piece_tokens) > 6:
                continue
            out.add(piece)
            if base.lower() not in piece.lower() and len(piece_tokens) <= 4:
                out.add(f"{base} {piece}")
    return {normalize_display(item) for item in out if normalize_display(item)}


def generate_rule_variants(alias_display: str) -> Set[str]:
    out: Set[str] = set()
    display = normalize_display(alias_display)
    tokens = list(tokenize(display))
    if not tokens:
        return out

    if "-" in display:
        out.add(display.replace("-", " "))
    if " " in display and "-" not in display and len(tokens) <= 3:
        out.add(display.replace(" ", "-"))

    if len(tokens) == 1:
        token = tokens[0]
        if token in GENERIC_RULE_VARIANT_BLOCKLIST:
            return out
        if len(token) >= 5 and token.endswith("s"):
            out.add(token[:-1])
        elif len(token) >= 4 and not token.endswith("s"):
            out.add(f"{token}s")
    return {normalize_display(item) for item in out if normalize_display(item)}


def ensure_canonical(
    class_key: str,
    canonical_name: str,
    source_type: str,
    source_origin: str,
    canonical_by_key: MutableMapping[Tuple[str, str], CanonicalEntry],
    canonical_by_id: MutableMapping[str, CanonicalEntry],
    canonical_id_counter: MutableMapping[str, int],
) -> CanonicalEntry:
    canonical_clean = normalize_display(canonical_name)
    canonical_norm_tokens = tokenize(canonical_clean)
    if not canonical_norm_tokens:
        raise ValueError("Canonical name cannot be empty after normalization.")

    lookup_key = (class_key, " ".join(canonical_norm_tokens))
    existing = canonical_by_key.get(lookup_key)
    if existing:
        return existing

    base_slug = f"{class_key}__{slugify(canonical_clean)}"
    suffix = canonical_id_counter.get(base_slug, 0)
    canonical_id_counter[base_slug] = suffix + 1
    canonical_id = base_slug if suffix == 0 else f"{base_slug}_{suffix+1}"

    entry = CanonicalEntry(
        canonical_id=canonical_id,
        class_key=class_key,
        canonical_name=canonical_clean,
        source_type=source_type,
        source_origin=source_origin,
    )
    canonical_by_key[lookup_key] = entry
    canonical_by_id[canonical_id] = entry
    return entry


def add_alias(
    class_key: str,
    canonical_id: str,
    alias_display: str,
    source_type: str,
    source_origin: str,
    status: str,
    aliases: List[AliasEntry],
    alias_token_index: MutableMapping[Tuple[str, Tuple[str, ...]], int],
) -> None:
    display = normalize_display(alias_display)
    tokens = tokenize(display)
    if not tokens:
        return

    key = (class_key, tokens)
    if key in alias_token_index:
        return

    alias_id = len(aliases)
    aliases.append(
        AliasEntry(
            alias_id=alias_id,
            canonical_id=canonical_id,
            class_key=class_key,
            alias_display=display,
            alias_tokens=tokens,
            source_type=source_type,
            source_origin=source_origin,
            status=status,
        )
    )
    alias_token_index[key] = alias_id


def ingest_workbook_terms(
    workbook_path: Path,
    canonical_by_key: MutableMapping[Tuple[str, str], CanonicalEntry],
    canonical_by_id: MutableMapping[str, CanonicalEntry],
    canonical_id_counter: MutableMapping[str, int],
    aliases: List[AliasEntry],
    alias_token_index: MutableMapping[Tuple[str, Tuple[str, ...]], int],
) -> openpyxl.Workbook:
    workbook = openpyxl.load_workbook(workbook_path)
    if WORKBOOK_SHEET_MAP not in workbook.sheetnames:
        raise ValueError(f"Workbook missing required sheet: {WORKBOOK_SHEET_MAP}")
    ws = workbook[WORKBOOK_SHEET_MAP]

    for row in range(2, ws.max_row + 1):
        allergen_name = normalize_display(str(ws.cell(row=row, column=1).value or ""))
        ingredient_cell = str(ws.cell(row=row, column=2).value or "").strip()
        if not allergen_name or not ingredient_cell:
            continue
        class_key = WORKBOOK_ALLERGEN_TO_CLASS.get(allergen_name)
        if not class_key:
            continue

        for part in split_slash_aliases(ingredient_cell):
            base = base_alias_without_parentheses(part)
            if not base:
                continue
            canonical = ensure_canonical(
                class_key=class_key,
                canonical_name=base,
                source_type="workbook_existing",
                source_origin=f"{WORKBOOK_SHEET_MAP}:{row}",
                canonical_by_key=canonical_by_key,
                canonical_by_id=canonical_by_id,
                canonical_id_counter=canonical_id_counter,
            )
            add_alias(
                class_key=class_key,
                canonical_id=canonical.canonical_id,
                alias_display=base,
                source_type="workbook_existing",
                source_origin=f"{WORKBOOK_SHEET_MAP}:{row}",
                status="approved",
                aliases=aliases,
                alias_token_index=alias_token_index,
            )
            for expanded in parenthetical_alias_expansions(part):
                add_alias(
                    class_key=class_key,
                    canonical_id=canonical.canonical_id,
                    alias_display=expanded,
                    source_type="workbook_parenthetical",
                    source_origin=f"{WORKBOOK_SHEET_MAP}:{row}",
                    status="approved",
                    aliases=aliases,
                    alias_token_index=alias_token_index,
                )
    return workbook


def ingest_manual_seeds(
    canonical_by_key: MutableMapping[Tuple[str, str], CanonicalEntry],
    canonical_by_id: MutableMapping[str, CanonicalEntry],
    canonical_id_counter: MutableMapping[str, int],
    aliases: List[AliasEntry],
    alias_token_index: MutableMapping[Tuple[str, Tuple[str, ...]], int],
) -> None:
    for class_key, canonical_map in MANUAL_CANONICAL_SEEDS.items():
        for canonical_name, alias_list in canonical_map.items():
            canonical = ensure_canonical(
                class_key=class_key,
                canonical_name=canonical_name,
                source_type="manual_seed",
                source_origin="manual_seed",
                canonical_by_key=canonical_by_key,
                canonical_by_id=canonical_by_id,
                canonical_id_counter=canonical_id_counter,
            )
            for alias in alias_list:
                add_alias(
                    class_key=class_key,
                    canonical_id=canonical.canonical_id,
                    alias_display=alias,
                    source_type="manual_seed",
                    source_origin="manual_seed",
                    status="approved",
                    aliases=aliases,
                    alias_token_index=alias_token_index,
                )


def ingest_rule_variants(
    aliases: List[AliasEntry],
    alias_token_index: MutableMapping[Tuple[str, Tuple[str, ...]], int],
) -> None:
    baseline_aliases = [alias for alias in aliases if alias.source_type in {"workbook_existing", "workbook_parenthetical", "manual_seed"}]
    for alias in baseline_aliases:
        for variant in generate_rule_variants(alias.alias_display):
            add_alias(
                class_key=alias.class_key,
                canonical_id=alias.canonical_id,
                alias_display=variant,
                source_type="rule_variant",
                source_origin=f"rule_from_alias:{alias.alias_id}",
                status="generated",
                aliases=aliases,
                alias_token_index=alias_token_index,
            )


def build_alias_trie(aliases: Sequence[AliasEntry]) -> Dict[str, object]:
    root: Dict[str, object] = {}
    for alias in aliases:
        node = root
        for token in alias.alias_tokens:
            child = node.get(token)
            if not isinstance(child, dict):
                child = {}
                node[token] = child
            node = child
        ids = node.get(TRIE_TERM_IDS_KEY)
        if not isinstance(ids, list):
            ids = []
            node[TRIE_TERM_IDS_KEY] = ids
        ids.append(alias.alias_id)
    return root


def match_aliases(tokens: Sequence[str], trie: Mapping[str, object]) -> Counter:
    out: Counter = Counter()
    count = len(tokens)
    for start in range(count):
        node = trie.get(tokens[start])
        if not isinstance(node, dict):
            continue

        ids = node.get(TRIE_TERM_IDS_KEY)
        if isinstance(ids, list):
            for alias_id in ids:
                out[alias_id] += 1

        idx = start + 1
        while idx < count:
            nxt = node.get(tokens[idx])
            if not isinstance(nxt, dict):
                break
            node = nxt
            ids = node.get(TRIE_TERM_IDS_KEY)
            if isinstance(ids, list):
                for alias_id in ids:
                    out[alias_id] += 1
            idx += 1
    return out


def extract_candidate_chunks(text: str) -> Set[Tuple[str, ...]]:
    normalized = normalize_ascii(text).lower()
    normalized = normalized.replace(":", " ")
    out: Set[Tuple[str, ...]] = set()
    for segment in CANDIDATE_SPLIT_RE.split(normalized):
        for raw_chunk in segment.split(","):
            chunk = " ".join(raw_chunk.strip().split())
            if not chunk:
                continue
            if any(ch.isdigit() for ch in chunk):
                continue
            if any(phrase in chunk for phrase in CANDIDATE_NOISE_PHRASES):
                continue
            tokens = tuple(TOKEN_RE.findall(chunk))
            if not tokens:
                continue
            if len(tokens) > 8:
                continue
            if sum(1 for tok in tokens if len(tok) <= 1) >= 2:
                continue
            informative = [tok for tok in tokens if tok not in STOPWORDS and len(tok) >= 3]
            if not informative:
                continue
            out.add(tokens)
    return out


def class_token_overlap(tokens: Sequence[str], class_key: str) -> bool:
    roots = CLASS_ROOT_TOKENS.get(class_key, set())
    if not roots:
        return False
    token_set = set(tokens)
    return any(root in token_set for root in roots)


def resolve_dataset_paths(args_dataset: Sequence[str], repo_root: Path) -> List[Path]:
    raw_paths = list(args_dataset) if args_dataset else list(DEFAULT_DATASET_FILES)
    out: List[Path] = []
    for raw in raw_paths:
        path = Path(raw)
        if not path.is_absolute():
            path = repo_root / path
        out.append(path)
    missing = [path for path in out if not path.exists()]
    if missing:
        lines = "\n".join(f"- {item}" for item in missing)
        raise FileNotFoundError(f"Missing dataset path(s):\n{lines}")
    return out


def resolve_optional_path(raw_path: str, repo_root: Path) -> Path | None:
    if not raw_path:
        return None
    path = Path(raw_path)
    if not path.is_absolute():
        path = repo_root / path
    return path


def load_alias_policy_rows(
    workbook: openpyxl.Workbook,
    denylist_csv: Path | None,
) -> Dict[Tuple[str, Tuple[str, ...]], AliasPolicyRow]:
    policy_rows: Dict[Tuple[str, Tuple[str, ...]], AliasPolicyRow] = {}

    def add_policy_row(
        class_raw: str,
        alias_raw: str,
        action_raw: str,
        reason_raw: str,
        source_origin: str,
    ) -> None:
        class_key = class_key_from_policy_label(class_raw)
        if not class_key:
            return
        tokens = tokenize(alias_raw)
        if not tokens:
            return
        action = normalize_display(action_raw).lower()
        if action not in VALID_ALIAS_POLICY_ACTIONS:
            return
        key = (class_key, tokens)
        policy_rows[key] = AliasPolicyRow(
            class_key=class_key,
            alias_tokens=tokens,
            action=action,
            reason=normalize_display(reason_raw),
            source_origin=source_origin,
        )

    if WORKBOOK_SHEET_ALIAS_POLICY in workbook.sheetnames:
        ws = workbook[WORKBOOK_SHEET_ALIAS_POLICY]
        header_map: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            value = normalize_display(str(ws.cell(row=1, column=col).value or "")).lower()
            if value:
                header_map[value] = col

        class_col = header_map.get("class") or header_map.get("class_key")
        alias_col = header_map.get("alias") or header_map.get("alias_norm") or header_map.get("candidate_alias")
        action_col = header_map.get("action")
        reason_col = header_map.get("reason")

        if class_col and alias_col and action_col:
            for row in range(2, ws.max_row + 1):
                class_raw = str(ws.cell(row=row, column=class_col).value or "")
                alias_raw = str(ws.cell(row=row, column=alias_col).value or "")
                action_raw = str(ws.cell(row=row, column=action_col).value or "")
                reason_raw = str(ws.cell(row=row, column=reason_col).value or "") if reason_col else ""
                add_policy_row(
                    class_raw=class_raw,
                    alias_raw=alias_raw,
                    action_raw=action_raw,
                    reason_raw=reason_raw,
                    source_origin=f"{WORKBOOK_SHEET_ALIAS_POLICY}:{row}",
                )

    if denylist_csv and denylist_csv.exists():
        with denylist_csv.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            for idx, row in enumerate(reader, start=2):
                class_raw = str(
                    row.get("class")
                    or row.get("class_key")
                    or row.get("allergen")
                    or row.get("scope")
                    or ""
                )
                alias_raw = str(
                    row.get("alias")
                    or row.get("alias_norm")
                    or row.get("candidate_alias")
                    or row.get("term")
                    or ""
                )
                action_raw = str(row.get("action") or row.get("status") or "")
                reason_raw = str(row.get("reason") or row.get("notes") or "")
                add_policy_row(
                    class_raw=class_raw,
                    alias_raw=alias_raw,
                    action_raw=action_raw,
                    reason_raw=reason_raw,
                    source_origin=f"{denylist_csv}:{idx}",
                )

    return policy_rows


def apply_policy_decisions(
    aliases: Sequence[AliasEntry],
    policy_rows: Mapping[Tuple[str, Tuple[str, ...]], AliasPolicyRow],
) -> Tuple[Set[int], Set[int], Dict[int, str], Dict[int, str], List[AliasActionRow]]:
    inactive_alias_ids: Set[int] = set()
    allow_alias_ids: Set[int] = set()
    status_override: Dict[int, str] = {}
    decision_reason: Dict[int, str] = {}
    actions: List[AliasActionRow] = []

    for alias in aliases:
        policy = resolve_policy_for_alias(alias.class_key, alias.alias_tokens, policy_rows)
        if policy and policy.action == "allow":
            allow_alias_ids.add(alias.alias_id)
            status_override[alias.alias_id] = "allowed_policy"
            reason = policy.reason or "Explicit allow policy."
            decision_reason[alias.alias_id] = reason
            actions.append(
                AliasActionRow(
                    alias_id=alias.alias_id,
                    class_key=alias.class_key,
                    canonical_id=alias.canonical_id,
                    alias_display=alias.alias_display,
                    alias_tokens=alias.alias_tokens,
                    source_type=alias.source_type,
                    source_origin=alias.source_origin,
                    status_before=alias.status,
                    status_after="allowed_policy",
                    action="allow_policy",
                    reason=reason,
                    rows_matched=0,
                    target_rows=0,
                    other_labeled_rows=0,
                    precision=0.0,
                    exclusivity=0.0,
                )
            )
            continue

        if policy and policy.action == "deny":
            inactive_alias_ids.add(alias.alias_id)
            status_override[alias.alias_id] = "denied_policy"
            reason = policy.reason or "Explicit deny policy."
            decision_reason[alias.alias_id] = reason
            actions.append(
                AliasActionRow(
                    alias_id=alias.alias_id,
                    class_key=alias.class_key,
                    canonical_id=alias.canonical_id,
                    alias_display=alias.alias_display,
                    alias_tokens=alias.alias_tokens,
                    source_type=alias.source_type,
                    source_origin=alias.source_origin,
                    status_before=alias.status,
                    status_after="denied_policy",
                    action="deny_policy",
                    reason=reason,
                    rows_matched=0,
                    target_rows=0,
                    other_labeled_rows=0,
                    precision=0.0,
                    exclusivity=0.0,
                )
            )
            continue

        if policy and policy.action == "review":
            allow_alias_ids.add(alias.alias_id)
            status_override[alias.alias_id] = "review_policy"
            reason = policy.reason or "Marked for manual review by policy."
            decision_reason[alias.alias_id] = reason
            actions.append(
                AliasActionRow(
                    alias_id=alias.alias_id,
                    class_key=alias.class_key,
                    canonical_id=alias.canonical_id,
                    alias_display=alias.alias_display,
                    alias_tokens=alias.alias_tokens,
                    source_type=alias.source_type,
                    source_origin=alias.source_origin,
                    status_before=alias.status,
                    status_after="review_policy",
                    action="review_policy",
                    reason=reason,
                    rows_matched=0,
                    target_rows=0,
                    other_labeled_rows=0,
                    precision=0.0,
                    exclusivity=0.0,
                )
            )
            continue

        if alias.alias_tokens in DEFAULT_AMBIGUOUS_ALIAS_TOKENS:
            inactive_alias_ids.add(alias.alias_id)
            status_override[alias.alias_id] = "denied_ambiguous"
            reason = "Auto-blocked generic ambiguous alias."
            decision_reason[alias.alias_id] = reason
            actions.append(
                AliasActionRow(
                    alias_id=alias.alias_id,
                    class_key=alias.class_key,
                    canonical_id=alias.canonical_id,
                    alias_display=alias.alias_display,
                    alias_tokens=alias.alias_tokens,
                    source_type=alias.source_type,
                    source_origin=alias.source_origin,
                    status_before=alias.status,
                    status_after="denied_ambiguous",
                    action="deny_ambiguous",
                    reason=reason,
                    rows_matched=0,
                    target_rows=0,
                    other_labeled_rows=0,
                    precision=0.0,
                    exclusivity=0.0,
                )
            )

    return inactive_alias_ids, allow_alias_ids, status_override, decision_reason, actions


def auto_demote_aliases(
    aliases: Sequence[AliasEntry],
    alias_stats: Mapping[int, AliasStats],
    inactive_alias_ids: Set[int],
    allow_alias_ids: Set[int],
    status_override: MutableMapping[int, str],
    decision_reason: MutableMapping[int, str],
    min_support: int,
    max_precision: float,
    max_exclusivity: float,
) -> List[AliasActionRow]:
    actions: List[AliasActionRow] = []
    for alias in aliases:
        alias_id = alias.alias_id
        if alias_id in inactive_alias_ids:
            continue
        if alias_id in allow_alias_ids:
            continue

        stat = alias_stats[alias_id]
        if stat.rows_matched < min_support:
            continue

        precision = alias_precision(stat)
        exclusivity = alias_exclusivity(stat)
        if precision > max_precision:
            continue
        if exclusivity > max_exclusivity:
            continue

        inactive_alias_ids.add(alias_id)
        status_override[alias_id] = "demoted_auto"
        reason = (
            f"Auto-demoted: support {stat.rows_matched}, "
            f"precision {precision:.3f}, exclusivity {exclusivity:.3f}."
        )
        decision_reason[alias_id] = reason
        actions.append(
            AliasActionRow(
                alias_id=alias_id,
                class_key=alias.class_key,
                canonical_id=alias.canonical_id,
                alias_display=alias.alias_display,
                alias_tokens=alias.alias_tokens,
                source_type=alias.source_type,
                source_origin=alias.source_origin,
                status_before=alias.status,
                status_after="demoted_auto",
                action="demote_auto",
                reason=reason,
                rows_matched=stat.rows_matched,
                target_rows=stat.target_rows,
                other_labeled_rows=stat.other_labeled_rows,
                precision=precision,
                exclusivity=exclusivity,
            )
        )
    return actions


def scan_datasets(
    dataset_paths: Sequence[Path],
    aliases: Sequence[AliasEntry],
    canonical_by_id: Mapping[str, CanonicalEntry],
    inactive_alias_ids: Set[int] | None = None,
) -> Tuple[
    int,
    Dict[int, AliasStats],
    Dict[str, CanonicalStats],
    Dict[str, ClassSummary],
    Dict[str, Counter],
    Dict[str, Counter],
    Dict[str, Counter],
    Dict[str, Counter],
    Dict[str, Counter],
]:
    inactive_set = inactive_alias_ids or set()
    active_aliases = [alias for alias in aliases if alias.alias_id not in inactive_set]
    trie = build_alias_trie(active_aliases)
    alias_stats: Dict[int, AliasStats] = {alias.alias_id: AliasStats() for alias in aliases}
    canonical_stats: Dict[str, CanonicalStats] = {cid: CanonicalStats() for cid in canonical_by_id}
    class_summary: Dict[str, ClassSummary] = {class_key: ClassSummary() for class_key in CLASS_ORDER}

    candidate_total: Dict[str, Counter] = {class_key: Counter() for class_key in CLASS_ORDER}
    candidate_target: Dict[str, Counter] = {class_key: Counter() for class_key in CLASS_ORDER}
    candidate_unmatched_target: Dict[str, Counter] = {class_key: Counter() for class_key in CLASS_ORDER}
    candidate_other_labeled: Dict[str, Counter] = {class_key: Counter() for class_key in CLASS_ORDER}
    candidate_unlabeled: Dict[str, Counter] = {class_key: Counter() for class_key in CLASS_ORDER}

    total_rows = 0

    for dataset_path in dataset_paths:
        with dataset_path.open("r", encoding="utf-8") as handle:
            for raw_line in handle:
                line = raw_line.strip()
                if not line:
                    continue
                payload = json.loads(line)
                if not isinstance(payload, dict):
                    continue

                total_rows += 1
                text = str(payload.get("text") or "")
                row_labels = {
                    allergen_key_from_label(str(item))
                    for item in (payload.get("allergens") or [])
                    if allergen_key_from_label(str(item))
                }
                wheat_proxy = "wheat" in row_labels

                for class_key in CLASS_ORDER:
                    if CLASS_SCOPE[class_key] == "big9":
                        if class_key in row_labels:
                            class_summary[class_key].labeled_rows += 1
                    else:
                        if wheat_proxy:
                            class_summary[class_key].proxy_rows += 1

                tokens = tokenize(text)
                matched_alias_counts = match_aliases(tokens, trie) if tokens else Counter()

                matched_classes_row: Set[str] = set()
                matched_canonicals_row: Set[str] = set()
                matched_canonical_target_row: Set[str] = set()

                for alias_id, mentions in matched_alias_counts.items():
                    alias = aliases[alias_id]
                    stat = alias_stats[alias_id]
                    stat.rows_matched += 1
                    stat.mentions_total += int(mentions)

                    target_hit = class_target_hit(alias.class_key, row_labels, wheat_proxy)
                    if target_hit:
                        stat.target_rows += 1
                    elif has_other_non_target_label(alias.class_key, row_labels):
                        stat.other_labeled_rows += 1

                    matched_classes_row.add(alias.class_key)
                    matched_canonicals_row.add(alias.canonical_id)
                    if target_hit:
                        matched_canonical_target_row.add(alias.canonical_id)

                    summary = class_summary[alias.class_key]
                    summary.alias_mentions_total += int(mentions)
                    summary.unique_alias_ids.add(alias_id)

                for class_key in matched_classes_row:
                    class_summary[class_key].rows_with_alias_any += 1
                    if CLASS_SCOPE[class_key] == "big9":
                        if class_key in row_labels:
                            class_summary[class_key].labeled_rows_with_alias += 1
                    else:
                        if wheat_proxy:
                            class_summary[class_key].proxy_rows_with_alias += 1

                for canonical_id in matched_canonicals_row:
                    canonical_stats[canonical_id].rows_matched += 1
                for canonical_id in matched_canonical_target_row:
                    canonical_stats[canonical_id].target_rows += 1
                for alias_id, mentions in matched_alias_counts.items():
                    canonical_id = aliases[alias_id].canonical_id
                    canonical_stats[canonical_id].mentions_total += int(mentions)

                chunks = extract_candidate_chunks(text)
                if chunks:
                    for class_key in CLASS_ORDER:
                        class_chunks = {chunk for chunk in chunks if class_token_overlap(chunk, class_key)}
                        if not class_chunks:
                            continue
                        candidate_total[class_key].update(class_chunks)

                        target_hit = class_target_hit(class_key, row_labels, wheat_proxy)
                        if target_hit:
                            candidate_target[class_key].update(class_chunks)
                            if class_key not in matched_classes_row:
                                candidate_unmatched_target[class_key].update(class_chunks)
                        elif has_other_non_target_label(class_key, row_labels):
                            candidate_other_labeled[class_key].update(class_chunks)
                        elif not row_labels:
                            candidate_unlabeled[class_key].update(class_chunks)

    return (
        total_rows,
        alias_stats,
        canonical_stats,
        class_summary,
        candidate_total,
        candidate_target,
        candidate_unmatched_target,
        candidate_other_labeled,
        candidate_unlabeled,
    )


def build_candidate_queue(
    candidate_total: Mapping[str, Counter],
    candidate_target: Mapping[str, Counter],
    candidate_unmatched_target: Mapping[str, Counter],
    candidate_other_labeled: Mapping[str, Counter],
    candidate_unlabeled: Mapping[str, Counter],
    alias_token_index: Mapping[Tuple[str, Tuple[str, ...]], int],
    policy_rows: Mapping[Tuple[str, Tuple[str, ...]], AliasPolicyRow],
    min_support: int,
    min_precision: float,
    min_exclusivity: float,
    min_unmatched_target_rows: int,
    max_review_per_class: int,
) -> List[CandidateRow]:
    queue: List[CandidateRow] = []

    for class_key in CLASS_ORDER:
        rows: List[CandidateRow] = []
        for token_tuple, unmatched_target in candidate_unmatched_target[class_key].items():
            unmatched_target_int = int(unmatched_target)
            if unmatched_target_int < min_unmatched_target_rows:
                continue
            support = int(candidate_total[class_key].get(token_tuple, 0))
            if support < min_support:
                continue
            if (class_key, token_tuple) in alias_token_index:
                continue
            if len(token_tuple) > 8:
                continue
            if not class_token_overlap(token_tuple, class_key):
                continue

            policy = resolve_policy_for_alias(class_key, token_tuple, policy_rows)
            if policy and policy.action == "deny":
                continue
            if not policy and token_tuple in DEFAULT_AMBIGUOUS_ALIAS_TOKENS:
                continue

            target = int(candidate_target[class_key].get(token_tuple, 0))
            other_labeled = int(candidate_other_labeled[class_key].get(token_tuple, 0))
            unlabeled = int(candidate_unlabeled[class_key].get(token_tuple, 0))
            precision = (float(target) / float(support)) if support else 0.0
            exclusivity = (float(target) / float(target + other_labeled)) if (target + other_labeled) else 0.0
            score = float(support) * precision * exclusivity
            display = title_from_tokens(token_tuple)

            if CLASS_SCOPE[class_key] == "big9":
                if target < max(8, int(min_support * 0.5)):
                    continue
                if precision >= 0.9 and exclusivity >= max(min_exclusivity, 0.75) and target >= max(20, min_support):
                    recommendation = "promote"
                    reason = "High support, precision, and exclusivity from uncovered positives."
                elif precision >= min_precision and exclusivity >= min_exclusivity:
                    recommendation = "review_high"
                    reason = "Strong precision/exclusivity; mined from unmatched positives."
                elif precision >= 0.6 and exclusivity >= 0.45 and unmatched_target_int >= min_unmatched_target_rows:
                    recommendation = "review_medium"
                    reason = "Moderate precision/exclusivity; verify before promotion."
                elif precision >= 0.45 and unmatched_target_int >= (2 * min_unmatched_target_rows):
                    recommendation = "review_low"
                    reason = "Frequent in uncovered positives but likely ambiguous."
                else:
                    continue
            else:
                # Cereals-with-gluten precision is proxy-only (wheat labels).
                if support >= max(40, min_support * 2) and unmatched_target_int >= min_unmatched_target_rows:
                    recommendation = "review_cereal_high"
                    reason = "High corpus support in uncovered wheat-proxy positives; verify gluten status."
                else:
                    recommendation = "review_cereal"
                    reason = "Candidate cereal term from uncovered proxy positives."

            rows.append(
                CandidateRow(
                    class_key=class_key,
                    candidate_display=display,
                    candidate_tokens=token_tuple,
                    support_rows=int(support),
                    target_rows=target,
                    unmatched_target_rows=unmatched_target_int,
                    other_labeled_rows=other_labeled,
                    unlabeled_rows=unlabeled,
                    precision=precision,
                    exclusivity=exclusivity,
                    score=score,
                    recommendation=recommendation,
                    reason=reason,
                )
            )

        if CLASS_SCOPE[class_key] == "big9":
            rows.sort(
                key=lambda item: (item.score, item.unmatched_target_rows, item.target_rows, item.support_rows, item.precision),
                reverse=True,
            )
        else:
            rows.sort(key=lambda item: (item.unmatched_target_rows, item.support_rows, item.exclusivity), reverse=True)
        queue.extend(rows[:max_review_per_class])

    queue.sort(
        key=lambda item: (
            CLASS_ORDER.index(item.class_key),
            -item.score,
            -item.unmatched_target_rows,
            -item.support_rows,
            -item.precision,
            item.candidate_display.lower(),
        )
    )
    return queue


def canonical_sort_key(item: CanonicalEntry, stats: CanonicalStats) -> Tuple[int, int, int, str]:
    return (
        CLASS_ORDER.index(item.class_key) if item.class_key in CLASS_ORDER else 999,
        -stats.target_rows,
        -stats.rows_matched,
        item.canonical_name.lower(),
    )


def alias_sort_key(item: AliasEntry, stats: AliasStats) -> Tuple[int, int, int, str]:
    return (
        CLASS_ORDER.index(item.class_key) if item.class_key in CLASS_ORDER else 999,
        -stats.target_rows,
        -stats.rows_matched,
        item.alias_display.lower(),
    )


def write_canonical_csv(
    output_path: Path,
    canonical_by_id: Mapping[str, CanonicalEntry],
    canonical_stats: Mapping[str, CanonicalStats],
    aliases: Sequence[AliasEntry],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    alias_count_by_canonical: Dict[str, int] = Counter(alias.canonical_id for alias in aliases)

    rows = sorted(canonical_by_id.values(), key=lambda item: canonical_sort_key(item, canonical_stats[item.canonical_id]))
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "canonical_id",
                "canonical_name",
                "source_type",
                "alias_count",
                "rows_matched",
                "target_rows",
                "target_precision_pct",
                "mentions_total",
                "source_origin",
            ]
        )
        for canonical in rows:
            stat = canonical_stats[canonical.canonical_id]
            precision = (100.0 * stat.target_rows / stat.rows_matched) if stat.rows_matched else 0.0
            writer.writerow(
                [
                    CLASS_DISPLAY[canonical.class_key],
                    CLASS_SCOPE[canonical.class_key],
                    canonical.canonical_id,
                    canonical.canonical_name,
                    canonical.source_type,
                    alias_count_by_canonical.get(canonical.canonical_id, 0),
                    stat.rows_matched,
                    stat.target_rows,
                    f"{precision:.2f}",
                    stat.mentions_total,
                    canonical.source_origin,
                ]
            )


def write_alias_csv(
    output_path: Path,
    aliases: Sequence[AliasEntry],
    alias_stats: Mapping[int, AliasStats],
    inactive_alias_ids: Set[int],
    status_override: Mapping[int, str],
    decision_reason: Mapping[int, str],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(aliases, key=lambda item: alias_sort_key(item, alias_stats[item.alias_id]))
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "canonical_id",
                "alias",
                "alias_norm",
                "source_type",
                "source_status",
                "effective_status",
                "active",
                "rows_matched",
                "target_rows",
                "other_labeled_rows",
                "target_precision_pct",
                "target_exclusivity_pct",
                "mentions_total",
                "source_origin",
                "decision_reason",
            ]
        )
        for alias in rows:
            stat = alias_stats[alias.alias_id]
            precision = 100.0 * alias_precision(stat)
            exclusivity = 100.0 * alias_exclusivity(stat)
            effective_status = status_override.get(alias.alias_id, alias.status)
            active = "no" if alias.alias_id in inactive_alias_ids else "yes"
            writer.writerow(
                [
                    CLASS_DISPLAY[alias.class_key],
                    CLASS_SCOPE[alias.class_key],
                    alias.canonical_id,
                    alias.alias_display,
                    " ".join(alias.alias_tokens),
                    alias.source_type,
                    alias.status,
                    effective_status,
                    active,
                    stat.rows_matched,
                    stat.target_rows,
                    stat.other_labeled_rows,
                    f"{precision:.2f}",
                    f"{exclusivity:.2f}",
                    stat.mentions_total,
                    alias.source_origin,
                    decision_reason.get(alias.alias_id, ""),
                ]
            )


def write_class_summary_csv(
    output_path: Path,
    class_summary: Mapping[str, ClassSummary],
    total_rows: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "total_rows_scanned",
                "labeled_rows",
                "labeled_rows_with_alias",
                "label_coverage_pct",
                "proxy_rows",
                "proxy_rows_with_alias",
                "proxy_coverage_pct",
                "rows_with_alias_any",
                "rows_with_alias_pct_total",
                "alias_mentions_total",
                "unique_aliases_matched",
            ]
        )
        for class_key in CLASS_ORDER:
            summary = class_summary[class_key]
            coverage = (100.0 * summary.labeled_rows_with_alias / summary.labeled_rows) if summary.labeled_rows else 0.0
            proxy_coverage = (100.0 * summary.proxy_rows_with_alias / summary.proxy_rows) if summary.proxy_rows else 0.0
            any_pct = (100.0 * summary.rows_with_alias_any / total_rows) if total_rows else 0.0
            writer.writerow(
                [
                    CLASS_DISPLAY[class_key],
                    CLASS_SCOPE[class_key],
                    total_rows,
                    summary.labeled_rows,
                    summary.labeled_rows_with_alias,
                    f"{coverage:.2f}",
                    summary.proxy_rows,
                    summary.proxy_rows_with_alias,
                    f"{proxy_coverage:.2f}",
                    summary.rows_with_alias_any,
                    f"{any_pct:.2f}",
                    summary.alias_mentions_total,
                    len(summary.unique_alias_ids or set()),
                ]
            )


def write_candidate_queue_csv(output_path: Path, queue: Sequence[CandidateRow]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "candidate_alias",
                "candidate_norm",
                "support_rows",
                "target_rows",
                "unmatched_target_rows",
                "other_labeled_rows",
                "unlabeled_rows",
                "target_precision_pct",
                "target_exclusivity_pct",
                "priority_score",
                "recommendation",
                "reason",
            ]
        )
        for item in queue:
            writer.writerow(
                [
                    CLASS_DISPLAY[item.class_key],
                    CLASS_SCOPE[item.class_key],
                    item.candidate_display,
                    " ".join(item.candidate_tokens),
                    item.support_rows,
                    item.target_rows,
                    item.unmatched_target_rows,
                    item.other_labeled_rows,
                    item.unlabeled_rows,
                    f"{100.0 * item.precision:.2f}",
                    f"{100.0 * item.exclusivity:.2f}",
                    f"{item.score:.4f}",
                    item.recommendation,
                    item.reason,
                ]
            )


def build_coverage_gaps(class_summary: Mapping[str, ClassSummary]) -> List[CoverageGapRow]:
    out: List[CoverageGapRow] = []
    for class_key in CLASS_ORDER:
        summary = class_summary[class_key]
        if CLASS_SCOPE[class_key] == "big9":
            target_rows = summary.labeled_rows
            covered_rows = summary.labeled_rows_with_alias
        else:
            target_rows = summary.proxy_rows
            covered_rows = summary.proxy_rows_with_alias
        uncovered = max(target_rows - covered_rows, 0)
        coverage = (100.0 * covered_rows / target_rows) if target_rows else 0.0
        out.append(
            CoverageGapRow(
                class_key=class_key,
                scope=CLASS_SCOPE[class_key],
                target_rows=target_rows,
                covered_rows=covered_rows,
                uncovered_rows=uncovered,
                coverage_pct=coverage,
            )
        )
    out.sort(key=lambda item: (-item.uncovered_rows, item.class_key))
    return out


def write_alias_actions_csv(output_path: Path, actions: Sequence[AliasActionRow]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    rows = sorted(
        actions,
        key=lambda item: (
            CLASS_ORDER.index(item.class_key) if item.class_key in CLASS_ORDER else 999,
            item.action,
            item.alias_display.lower(),
        ),
    )
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "canonical_id",
                "alias_id",
                "alias",
                "alias_norm",
                "source_type",
                "source_origin",
                "status_before",
                "status_after",
                "action",
                "rows_matched",
                "target_rows",
                "other_labeled_rows",
                "target_precision_pct",
                "target_exclusivity_pct",
                "reason",
            ]
        )
        for item in rows:
            writer.writerow(
                [
                    CLASS_DISPLAY[item.class_key],
                    CLASS_SCOPE[item.class_key],
                    item.canonical_id,
                    item.alias_id,
                    item.alias_display,
                    " ".join(item.alias_tokens),
                    item.source_type,
                    item.source_origin,
                    item.status_before,
                    item.status_after,
                    item.action,
                    item.rows_matched,
                    item.target_rows,
                    item.other_labeled_rows,
                    f"{100.0 * item.precision:.2f}",
                    f"{100.0 * item.exclusivity:.2f}",
                    item.reason,
                ]
            )


def write_coverage_gap_csv(output_path: Path, gaps: Sequence[CoverageGapRow]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "class",
                "class_scope",
                "target_rows",
                "covered_rows",
                "uncovered_rows",
                "coverage_pct",
            ]
        )
        for item in gaps:
            writer.writerow(
                [
                    CLASS_DISPLAY[item.class_key],
                    item.scope,
                    item.target_rows,
                    item.covered_rows,
                    item.uncovered_rows,
                    f"{item.coverage_pct:.2f}",
                ]
            )


def rebuild_sheet(workbook: openpyxl.Workbook, name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if name in workbook.sheetnames:
        old = workbook[name]
        workbook.remove(old)
    return workbook.create_sheet(name)


def populate_workbook_sheets(
    workbook: openpyxl.Workbook,
    canonical_by_id: Mapping[str, CanonicalEntry],
    canonical_stats: Mapping[str, CanonicalStats],
    aliases: Sequence[AliasEntry],
    alias_stats: Mapping[int, AliasStats],
    inactive_alias_ids: Set[int],
    status_override: Mapping[int, str],
    decision_reason: Mapping[int, str],
    class_summary: Mapping[str, ClassSummary],
    candidate_queue: Sequence[CandidateRow],
    alias_actions: Sequence[AliasActionRow],
    coverage_gaps: Sequence[CoverageGapRow],
    policy_rows: Mapping[Tuple[str, Tuple[str, ...]], AliasPolicyRow],
    total_rows: int,
) -> None:
    ws_canonical = rebuild_sheet(workbook, "Lexicon Canonical")
    ws_aliases = rebuild_sheet(workbook, "Lexicon Aliases")
    ws_summary = rebuild_sheet(workbook, "Lexicon Class Summary")
    ws_queue = rebuild_sheet(workbook, "Lexicon Candidate Queue")
    ws_actions = rebuild_sheet(workbook, "Lexicon Alias Actions")
    ws_gaps = rebuild_sheet(workbook, "Lexicon Coverage Gaps")
    ws_policy = rebuild_sheet(workbook, WORKBOOK_SHEET_ALIAS_POLICY)

    ws_canonical.append(
        [
            "Class",
            "Scope",
            "Canonical ID",
            "Canonical Name",
            "Source Type",
            "Rows Matched",
            "Target Rows",
            "Target Precision %",
            "Mentions Total",
            "Source Origin",
        ]
    )
    for canonical in sorted(canonical_by_id.values(), key=lambda item: canonical_sort_key(item, canonical_stats[item.canonical_id])):
        stat = canonical_stats[canonical.canonical_id]
        precision = (100.0 * stat.target_rows / stat.rows_matched) if stat.rows_matched else 0.0
        ws_canonical.append(
            [
                CLASS_DISPLAY[canonical.class_key],
                CLASS_SCOPE[canonical.class_key],
                canonical.canonical_id,
                canonical.canonical_name,
                canonical.source_type,
                stat.rows_matched,
                stat.target_rows,
                round(precision, 2),
                stat.mentions_total,
                canonical.source_origin,
            ]
        )
    ws_canonical.freeze_panes = "A2"

    ws_aliases.append(
        [
            "Class",
            "Scope",
            "Canonical ID",
            "Alias",
            "Alias Norm",
            "Source Type",
            "Source Status",
            "Effective Status",
            "Active",
            "Rows Matched",
            "Target Rows",
            "Other Labeled Rows",
            "Target Precision %",
            "Target Exclusivity %",
            "Mentions Total",
            "Source Origin",
            "Decision Reason",
        ]
    )
    for alias in sorted(aliases, key=lambda item: alias_sort_key(item, alias_stats[item.alias_id])):
        stat = alias_stats[alias.alias_id]
        precision = 100.0 * alias_precision(stat)
        exclusivity = 100.0 * alias_exclusivity(stat)
        effective_status = status_override.get(alias.alias_id, alias.status)
        active = "no" if alias.alias_id in inactive_alias_ids else "yes"
        ws_aliases.append(
            [
                CLASS_DISPLAY[alias.class_key],
                CLASS_SCOPE[alias.class_key],
                alias.canonical_id,
                alias.alias_display,
                " ".join(alias.alias_tokens),
                alias.source_type,
                alias.status,
                effective_status,
                active,
                stat.rows_matched,
                stat.target_rows,
                stat.other_labeled_rows,
                round(precision, 2),
                round(exclusivity, 2),
                stat.mentions_total,
                alias.source_origin,
                decision_reason.get(alias.alias_id, ""),
            ]
        )
    ws_aliases.freeze_panes = "A2"

    ws_summary.append(
        [
            "Class",
            "Scope",
            "Total Rows Scanned",
            "Labeled Rows",
            "Labeled Rows With Alias",
            "Label Coverage %",
            "Proxy Rows",
            "Proxy Rows With Alias",
            "Proxy Coverage %",
            "Rows With Alias Any",
            "Rows With Alias % Total",
            "Alias Mentions Total",
            "Unique Aliases Matched",
        ]
    )
    for class_key in CLASS_ORDER:
        summary = class_summary[class_key]
        coverage = (100.0 * summary.labeled_rows_with_alias / summary.labeled_rows) if summary.labeled_rows else 0.0
        proxy_coverage = (100.0 * summary.proxy_rows_with_alias / summary.proxy_rows) if summary.proxy_rows else 0.0
        any_pct = (100.0 * summary.rows_with_alias_any / total_rows) if total_rows else 0.0
        ws_summary.append(
            [
                CLASS_DISPLAY[class_key],
                CLASS_SCOPE[class_key],
                total_rows,
                summary.labeled_rows,
                summary.labeled_rows_with_alias,
                round(coverage, 2),
                summary.proxy_rows,
                summary.proxy_rows_with_alias,
                round(proxy_coverage, 2),
                summary.rows_with_alias_any,
                round(any_pct, 2),
                summary.alias_mentions_total,
                len(summary.unique_alias_ids or set()),
            ]
        )
    ws_summary.freeze_panes = "A2"

    ws_queue.append(
        [
            "Class",
            "Scope",
            "Candidate Alias",
            "Candidate Norm",
            "Support Rows",
            "Target Rows",
            "Unmatched Target Rows",
            "Other Labeled Rows",
            "Unlabeled Rows",
            "Target Precision %",
            "Target Exclusivity %",
            "Priority Score",
            "Recommendation",
            "Reason",
        ]
    )
    for item in candidate_queue:
        ws_queue.append(
            [
                CLASS_DISPLAY[item.class_key],
                CLASS_SCOPE[item.class_key],
                item.candidate_display,
                " ".join(item.candidate_tokens),
                item.support_rows,
                item.target_rows,
                item.unmatched_target_rows,
                item.other_labeled_rows,
                item.unlabeled_rows,
                round(100.0 * item.precision, 2),
                round(100.0 * item.exclusivity, 2),
                round(item.score, 4),
                item.recommendation,
                item.reason,
            ]
        )
    ws_queue.freeze_panes = "A2"

    ws_actions.append(
        [
            "Class",
            "Scope",
            "Canonical ID",
            "Alias ID",
            "Alias",
            "Alias Norm",
            "Source Type",
            "Source Origin",
            "Status Before",
            "Status After",
            "Action",
            "Rows Matched",
            "Target Rows",
            "Other Labeled Rows",
            "Target Precision %",
            "Target Exclusivity %",
            "Reason",
        ]
    )
    for item in sorted(
        alias_actions,
        key=lambda row: (
            CLASS_ORDER.index(row.class_key) if row.class_key in CLASS_ORDER else 999,
            row.action,
            row.alias_display.lower(),
        ),
    ):
        ws_actions.append(
            [
                CLASS_DISPLAY[item.class_key],
                CLASS_SCOPE[item.class_key],
                item.canonical_id,
                item.alias_id,
                item.alias_display,
                " ".join(item.alias_tokens),
                item.source_type,
                item.source_origin,
                item.status_before,
                item.status_after,
                item.action,
                item.rows_matched,
                item.target_rows,
                item.other_labeled_rows,
                round(100.0 * item.precision, 2),
                round(100.0 * item.exclusivity, 2),
                item.reason,
            ]
        )
    ws_actions.freeze_panes = "A2"

    ws_gaps.append(
        [
            "Class",
            "Scope",
            "Target Rows",
            "Covered Rows",
            "Uncovered Rows",
            "Coverage %",
        ]
    )
    for gap in coverage_gaps:
        ws_gaps.append(
            [
                CLASS_DISPLAY[gap.class_key],
                gap.scope,
                gap.target_rows,
                gap.covered_rows,
                gap.uncovered_rows,
                round(gap.coverage_pct, 2),
            ]
        )
    ws_gaps.freeze_panes = "A2"

    ws_policy.append(
        [
            "Class",
            "Alias",
            "Alias Norm",
            "Action",
            "Reason",
            "Source Origin",
        ]
    )
    policy_sorted = sorted(
        policy_rows.values(),
        key=lambda row: (
            CLASS_ORDER.index(row.class_key) if row.class_key in CLASS_ORDER else 999,
            row.action,
            " ".join(row.alias_tokens),
        ),
    )
    for policy in policy_sorted:
        class_label = CLASS_DISPLAY.get(policy.class_key, "All")
        if policy.class_key == POLICY_CLASS_ANY:
            class_label = "*"
        alias_norm = " ".join(policy.alias_tokens)
        ws_policy.append(
            [
                class_label,
                alias_norm,
                alias_norm,
                policy.action,
                policy.reason,
                policy.source_origin,
            ]
        )
    ws_policy.freeze_panes = "A2"


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[2]

    workbook_input = Path(args.workbook_input)
    if not workbook_input.exists():
        print(f"Workbook not found: {workbook_input}")
        return 1
    dataset_paths = resolve_dataset_paths(args.dataset, repo_root)
    denylist_csv = resolve_optional_path(str(args.denylist_csv or ""), repo_root)
    if args.denylist_csv and denylist_csv and not denylist_csv.exists():
        print(f"Denylist CSV not found: {denylist_csv}")
        return 1

    canonical_by_key: Dict[Tuple[str, str], CanonicalEntry] = {}
    canonical_by_id: Dict[str, CanonicalEntry] = {}
    canonical_id_counter: Dict[str, int] = {}
    aliases: List[AliasEntry] = []
    alias_token_index: Dict[Tuple[str, Tuple[str, ...]], int] = {}

    workbook = ingest_workbook_terms(
        workbook_path=workbook_input,
        canonical_by_key=canonical_by_key,
        canonical_by_id=canonical_by_id,
        canonical_id_counter=canonical_id_counter,
        aliases=aliases,
        alias_token_index=alias_token_index,
    )
    ingest_manual_seeds(
        canonical_by_key=canonical_by_key,
        canonical_by_id=canonical_by_id,
        canonical_id_counter=canonical_id_counter,
        aliases=aliases,
        alias_token_index=alias_token_index,
    )
    ingest_rule_variants(aliases=aliases, alias_token_index=alias_token_index)

    policy_rows = load_alias_policy_rows(workbook=workbook, denylist_csv=denylist_csv)
    inactive_alias_ids, allow_alias_ids, status_override, decision_reason, alias_actions = apply_policy_decisions(
        aliases=aliases,
        policy_rows=policy_rows,
    )

    (
        _total_rows_pass1,
        alias_stats_pass1,
        _canonical_stats_pass1,
        _class_summary_pass1,
        _candidate_total_pass1,
        _candidate_target_pass1,
        _candidate_unmatched_pass1,
        _candidate_other_labeled_pass1,
        _candidate_unlabeled_pass1,
    ) = scan_datasets(
        dataset_paths=dataset_paths,
        aliases=aliases,
        canonical_by_id=canonical_by_id,
        inactive_alias_ids=inactive_alias_ids,
    )

    alias_actions.extend(
        auto_demote_aliases(
            aliases=aliases,
            alias_stats=alias_stats_pass1,
            inactive_alias_ids=inactive_alias_ids,
            allow_alias_ids=allow_alias_ids,
            status_override=status_override,
            decision_reason=decision_reason,
            min_support=int(args.auto_demote_min_support),
            max_precision=float(args.auto_demote_max_precision),
            max_exclusivity=float(args.auto_demote_max_exclusivity),
        )
    )

    (
        total_rows,
        alias_stats,
        canonical_stats,
        class_summary,
        candidate_total,
        candidate_target,
        candidate_unmatched_target,
        candidate_other_labeled,
        candidate_unlabeled,
    ) = scan_datasets(
        dataset_paths=dataset_paths,
        aliases=aliases,
        canonical_by_id=canonical_by_id,
        inactive_alias_ids=inactive_alias_ids,
    )

    candidate_queue = build_candidate_queue(
        candidate_total=candidate_total,
        candidate_target=candidate_target,
        candidate_unmatched_target=candidate_unmatched_target,
        candidate_other_labeled=candidate_other_labeled,
        candidate_unlabeled=candidate_unlabeled,
        alias_token_index=alias_token_index,
        policy_rows=policy_rows,
        min_support=int(args.min_candidate_support),
        min_precision=float(args.min_candidate_precision),
        min_exclusivity=float(args.candidate_min_exclusivity),
        min_unmatched_target_rows=int(args.candidate_min_unmatched_target_rows),
        max_review_per_class=int(args.max_review_per_class),
    )
    coverage_gaps = build_coverage_gaps(class_summary)

    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = repo_root / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    canonical_csv = output_dir / "allergen_lexicon_canonical.csv"
    alias_csv = output_dir / "allergen_lexicon_aliases.csv"
    class_summary_csv = output_dir / "allergen_lexicon_class_summary.csv"
    candidate_csv = output_dir / "allergen_lexicon_candidate_queue.csv"
    alias_actions_csv = output_dir / "allergen_lexicon_alias_actions.csv"
    coverage_gaps_csv = output_dir / "allergen_lexicon_coverage_gaps.csv"

    write_canonical_csv(canonical_csv, canonical_by_id, canonical_stats, aliases)
    write_alias_csv(
        alias_csv,
        aliases,
        alias_stats,
        inactive_alias_ids=inactive_alias_ids,
        status_override=status_override,
        decision_reason=decision_reason,
    )
    write_class_summary_csv(class_summary_csv, class_summary, total_rows)
    write_candidate_queue_csv(candidate_csv, candidate_queue)
    write_alias_actions_csv(alias_actions_csv, alias_actions)
    write_coverage_gap_csv(coverage_gaps_csv, coverage_gaps)

    populate_workbook_sheets(
        workbook=workbook,
        canonical_by_id=canonical_by_id,
        canonical_stats=canonical_stats,
        aliases=aliases,
        alias_stats=alias_stats,
        inactive_alias_ids=inactive_alias_ids,
        status_override=status_override,
        decision_reason=decision_reason,
        class_summary=class_summary,
        candidate_queue=candidate_queue,
        alias_actions=alias_actions,
        coverage_gaps=coverage_gaps,
        policy_rows=policy_rows,
        total_rows=total_rows,
    )

    workbook_output = Path(args.workbook_output)
    workbook_output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_output)

    active_alias_count = len(aliases) - len(inactive_alias_ids)
    auto_demoted_count = sum(1 for item in alias_actions if item.action == "demote_auto")

    print(f"Workbook input: {workbook_input}")
    print(f"Workbook output: {workbook_output}")
    print(f"Datasets scanned: {len(dataset_paths)}")
    print(f"Total rows scanned: {total_rows}")
    print(f"Canonical terms: {len(canonical_by_id)}")
    print(f"Alias terms: {len(aliases)}")
    print(f"Active aliases after policy+demotion: {active_alias_count}")
    print(f"Auto-demoted aliases: {auto_demoted_count}")
    print(f"Policy rows loaded: {len(policy_rows)}")
    print(f"Candidate review rows: {len(candidate_queue)}")
    print(f"Canonical CSV: {canonical_csv}")
    print(f"Alias CSV: {alias_csv}")
    print(f"Class summary CSV: {class_summary_csv}")
    print(f"Candidate queue CSV: {candidate_csv}")
    print(f"Alias actions CSV: {alias_actions_csv}")
    print(f"Coverage gaps CSV: {coverage_gaps_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
