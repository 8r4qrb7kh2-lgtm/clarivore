#!/usr/bin/env python3
"""Expand allergen ingredient terms and compute dataset frequencies."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Mapping, MutableMapping, Sequence, Set, Tuple

import openpyxl


ALLERGEN_KEY_BY_WORKBOOK: Mapping[str, str] = {
    "Milk": "milk",
    "Egg": "egg",
    "Peanut": "peanut",
    "Tree Nut": "tree nut",
    "Wheat": "wheat",
    "Soy": "soy",
    "Fish": "fish",
    "Crustacean Shellfish": "shellfish",
    "Sesame": "sesame",
}
WORKBOOK_ALLERGEN_BY_KEY: Mapping[str, str] = {value: key for key, value in ALLERGEN_KEY_BY_WORKBOOK.items()}
ALLERGEN_ORDER: Sequence[str] = (
    "milk",
    "egg",
    "peanut",
    "tree nut",
    "wheat",
    "soy",
    "fish",
    "shellfish",
    "sesame",
)

DEFAULT_DATASET_FILES: Sequence[str] = (
    "ml/data/processed/usda_only_train.jsonl",
    "ml/data/processed/usda_only_val.jsonl",
    "ml/data/processed/usda_only_holdout.jsonl",
    "ml/data/processed/openfoodfacts_targeted_examples.jsonl",
)

MANUAL_EXPANSIONS: Mapping[str, Sequence[str]] = {
    "milk": (
        "Buttermilk",
        "Cultured Milk",
        "Condensed Milk",
        "Evaporated Milk",
        "Nonfat Dry Milk",
        "Skim Milk Powder",
        "Milk Powder",
        "Milkfat",
        "Milk Solids",
        "Milk Solids Nonfat",
        "Milk Permeate",
        "Whey Protein Concentrate",
        "Whey Protein Isolate",
        "Whey Protein Hydrolysate",
        "Whey Powder",
        "Sweet Whey",
        "Sodium Caseinate",
        "Calcium Caseinate",
        "Potassium Caseinate",
    ),
    "egg": (
        "Egg White",
        "Egg Whites",
        "Egg Yolk",
        "Egg Yolks",
        "Dried Egg White",
        "Dried Egg Yolk",
        "Whole Egg Solids",
        "Liquid Egg",
        "Pasteurized Eggs",
        "Egg Protein",
        "Egg Albumin",
        "Egg Powder",
    ),
    "peanut": (
        "Peanut",
        "Peanuts",
        "Peanut Oil",
        "Dry Roasted Peanuts",
        "Groundnut Oil",
        "Arachis Hypogaea",
        "Peanut Pieces",
        "Peanut Paste",
        "Peanut Extract",
    ),
    "tree nut": (
        "Almond Flour",
        "Almond Meal",
        "Almond Milk",
        "Cashew Milk",
        "Walnut Pieces",
        "Pecan Pieces",
        "Pistachio Paste",
        "Hazelnut Spread",
        "Macadamia Oil",
        "Coconut Milk",
        "Coconut Cream",
        "Coconut Flour",
        "Chestnut Flour",
        "Pine Nut Oil",
    ),
    "wheat": (
        "Wheat Flour",
        "Whole Wheat Flour",
        "Enriched Wheat Flour",
        "Durum Wheat",
        "Durum Wheat Semolina",
        "Wheat Starch",
        "Wheat Gluten",
        "Gluten Flour",
        "Vital Gluten",
        "Wheat Germ",
        "Wheat Bran",
        "Semolina Flour",
    ),
    "soy": (
        "Soybeans",
        "Soy Bean",
        "Soybean Meal",
        "Soybean Lecithin",
        "Soy Milk Powder",
        "Soymilk Powder",
        "Hydrolyzed Soy Protein",
        "Soy Protein Isolate",
        "Soy Protein Concentrate",
        "Soy Protein Hydrolysate",
        "Defatted Soy Flour",
        "Soy Isolate",
        "Soy Grits",
        "Soy Protein Powder",
        "Soybean Protein",
    ),
    "fish": (
        "Sardine",
        "Sardines",
        "Anchovies",
        "Mackerel",
        "Herring",
        "Anchoveta",
        "Menhaden",
        "Bonito",
        "Eel",
        "Fish Extract",
        "Fish Protein",
        "Fish Powder",
        "Pollock Fillet",
    ),
    "shellfish": (
        "Clam",
        "Clams",
        "Mussel",
        "Mussels",
        "Oyster",
        "Oysters",
        "Scallop",
        "Scallops",
        "Shrimp Powder",
        "Shrimp Extract",
        "Crab Extract",
        "Lobster Extract",
        "Crab Meat",
        "Prawn Extract",
        "Krill Meal",
    ),
    "sesame": (
        "Sesame Seeds",
        "Toasted Sesame Oil",
        "Black Sesame",
        "White Sesame",
        "Sesame Meal",
        "Sesame Butter",
        "Sesame Tahini",
        "Tahini Paste",
        "Sesame Protein",
        "Sesame Powder",
        "Sesamum",
        "Til Seeds",
        "Benne Flour",
    ),
}

TERM_IDS_KEY = "__term_ids__"
TEXT_TOKEN_RE = re.compile(r"[a-z0-9]+")
PAREN_RE = re.compile(r"\(([^()]*)\)")
PAREN_SPLIT_RE = re.compile(r"[,/;]")
PAREN_STRIP_PREFIX_RE = re.compile(
    r"^(all forms:?|all types:?|including|includes?|incl\.?|e\.g\.?|such as|if [a-z -]+derived)\s+",
    re.IGNORECASE,
)
DISPLAY_CLEAN_RE = re.compile(r"\s+")

GENERIC_SINGLE_WORD_EXPANSION_BLOCKLIST: Set[str] = {
    "acid",
    "blend",
    "extract",
    "fat",
    "flavor",
    "flavour",
    "flour",
    "meal",
    "milk",
    "nut",
    "nuts",
    "oil",
    "paste",
    "powder",
    "protein",
    "sauce",
    "seed",
    "seeds",
    "solids",
    "starch",
    "wheat",
    "egg",
    "soy",
    "fish",
    "shellfish",
    "sesame",
}


@dataclass(frozen=True)
class TermEntry:
    term_id: int
    allergen_key: str
    allergen_name: str
    term_display: str
    term_tokens: Tuple[str, ...]
    source_type: str
    source_origin: str


@dataclass
class TermStats:
    rows_matched: int = 0
    rows_with_target_allergen: int = 0
    mentions_total: int = 0


@dataclass
class AllergenSummary:
    labeled_rows: int = 0
    labeled_rows_with_term: int = 0
    term_mentions: int = 0
    unique_term_ids: Set[int] | None = None

    def __post_init__(self) -> None:
        if self.unique_term_ids is None:
            self.unique_term_ids = set()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Expand allergen ingredient names and compute term frequencies.")
    parser.add_argument(
        "--workbook-input",
        default="/Users/mattdavis/Desktop/allergen_ingredient_database.xlsx",
        help="Path to the source allergen ingredient workbook.",
    )
    parser.add_argument(
        "--workbook-output",
        default="/Users/mattdavis/Desktop/allergen_ingredient_database_enriched.xlsx",
        help="Path to save the enriched workbook.",
    )
    parser.add_argument(
        "--dataset",
        action="append",
        default=[],
        help="JSONL dataset path (repeatable). Defaults to USDA/OpenFoodFacts processed files.",
    )
    parser.add_argument(
        "--csv-output-dir",
        default="ml/data/analysis",
        help="Directory for CSV analysis outputs.",
    )
    parser.add_argument(
        "--min-added-target-rows",
        type=int,
        default=10,
        help="Minimum target-allergen row matches required before adding a new term to workbook map.",
    )
    parser.add_argument(
        "--min-added-precision",
        type=float,
        default=0.70,
        help="Minimum precision (target matches / matched rows) for added workbook terms.",
    )
    parser.add_argument(
        "--include-parenthetical-additions",
        action="store_true",
        help="If set, allow parenthetical auto-expansion terms to be appended to workbook map.",
    )
    return parser.parse_args()


def normalize_ascii(value: str) -> str:
    return unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii")


def tokenize_text(value: str) -> List[str]:
    safe = normalize_ascii(value).lower()
    return TEXT_TOKEN_RE.findall(safe)


def normalize_display_term(value: str) -> str:
    safe = normalize_ascii(value or "")
    safe = safe.replace("®", "").replace("™", "")
    safe = safe.replace("’", "'")
    safe = DISPLAY_CLEAN_RE.sub(" ", safe).strip()
    safe = safe.strip(",;:.")
    return safe


def clean_parenthetical_fragment(fragment: str) -> str:
    safe = normalize_display_term(fragment)
    safe = PAREN_STRIP_PREFIX_RE.sub("", safe).strip()
    safe = safe.replace("etc.", "").replace("etc", "").strip(" ,;:.")
    return safe


def allergen_key_from_label(value: str) -> str:
    safe = (value or "").strip().lower().replace("-", " ")
    safe = re.sub(r"\s+", " ", safe)
    mapping = {
        "milk": "milk",
        "egg": "egg",
        "peanut": "peanut",
        "tree nut": "tree nut",
        "tree nuts": "tree nut",
        "wheat": "wheat",
        "soy": "soy",
        "fish": "fish",
        "shellfish": "shellfish",
        "crustacean shellfish": "shellfish",
        "sesame": "sesame",
    }
    return mapping.get(safe, "")


def split_slash_aliases(value: str) -> List[str]:
    parts = [part.strip() for part in re.split(r"\s*/\s*", value or "") if part.strip()]
    return parts


def base_alias_without_parentheses(alias: str) -> str:
    base = PAREN_RE.sub("", alias)
    base = normalize_display_term(base)
    return base


def parenthetical_expansion_candidates(alias: str) -> Set[str]:
    out: Set[str] = set()
    base = base_alias_without_parentheses(alias)
    base_tokens = tokenize_text(base)
    if not base_tokens:
        return out

    for match in PAREN_RE.finditer(alias):
        raw_inner = match.group(1)
        for raw_piece in PAREN_SPLIT_RE.split(raw_inner):
            piece = clean_parenthetical_fragment(raw_piece)
            if not piece:
                continue
            piece_tokens = tokenize_text(piece)
            if not piece_tokens or len(piece_tokens) > 5:
                continue
            lower_piece = piece.lower()
            if lower_piece in {"all forms", "all types", "cross reactive"}:
                continue

            out.add(piece)

            base_lower = base.lower()
            if base_lower not in lower_piece:
                out.add(f"{base} {piece}")
                out.add(f"{piece} {base}")
    return {normalize_display_term(item) for item in out if normalize_display_term(item)}


def add_term_entry(
    entries: List[TermEntry],
    seen_tokens: MutableMapping[Tuple[str, Tuple[str, ...]], int],
    allergen_key: str,
    term_display: str,
    source_type: str,
    source_origin: str,
) -> None:
    display = normalize_display_term(term_display)
    tokens = tuple(tokenize_text(display))
    if not tokens:
        return

    key = (allergen_key, tokens)
    if key in seen_tokens:
        return

    term_id = len(entries)
    entries.append(
        TermEntry(
            term_id=term_id,
            allergen_key=allergen_key,
            allergen_name=WORKBOOK_ALLERGEN_BY_KEY[allergen_key],
            term_display=display,
            term_tokens=tokens,
            source_type=source_type,
            source_origin=source_origin,
        )
    )
    seen_tokens[key] = term_id


def build_term_inventory(workbook_path: Path) -> Tuple[openpyxl.Workbook, List[TermEntry], Set[Tuple[str, Tuple[str, ...]]]]:
    workbook = openpyxl.load_workbook(workbook_path)
    if "Allergen-Ingredient Map" not in workbook.sheetnames:
        raise ValueError("Workbook is missing required sheet: Allergen-Ingredient Map")
    ws_map = workbook["Allergen-Ingredient Map"]

    entries: List[TermEntry] = []
    seen_tokens: Dict[Tuple[str, Tuple[str, ...]], int] = {}
    existing_token_keys: Set[Tuple[str, Tuple[str, ...]]] = set()

    for row in range(2, ws_map.max_row + 1):
        workbook_allergen = normalize_display_term(str(ws_map.cell(row=row, column=1).value or ""))
        raw_alias = str(ws_map.cell(row=row, column=2).value or "").strip()
        if not workbook_allergen or not raw_alias:
            continue

        allergen_key = ALLERGEN_KEY_BY_WORKBOOK.get(workbook_allergen)
        if not allergen_key:
            continue

        for alias_part in split_slash_aliases(raw_alias):
            base_term = base_alias_without_parentheses(alias_part)
            if base_term:
                add_term_entry(
                    entries,
                    seen_tokens,
                    allergen_key=allergen_key,
                    term_display=base_term,
                    source_type="existing",
                    source_origin=alias_part,
                )
                existing_token_keys.add((allergen_key, tuple(tokenize_text(base_term))))

            for expanded_term in parenthetical_expansion_candidates(alias_part):
                add_term_entry(
                    entries,
                    seen_tokens,
                    allergen_key=allergen_key,
                    term_display=expanded_term,
                    source_type="expanded_parenthetical",
                    source_origin=alias_part,
                )

    for allergen_key, terms in MANUAL_EXPANSIONS.items():
        if allergen_key not in WORKBOOK_ALLERGEN_BY_KEY:
            continue
        for term in terms:
            add_term_entry(
                entries,
                seen_tokens,
                allergen_key=allergen_key,
                term_display=term,
                source_type="expanded_manual",
                source_origin="manual_expansion_seed",
            )

    return workbook, entries, existing_token_keys


def build_term_trie(entries: Sequence[TermEntry]) -> Dict[str, object]:
    root: Dict[str, object] = {}
    for entry in entries:
        node: Dict[str, object] = root
        for token in entry.term_tokens:
            child = node.get(token)
            if not isinstance(child, dict):
                child = {}
                node[token] = child
            node = child
        ids = node.get(TERM_IDS_KEY)
        if not isinstance(ids, list):
            ids = []
            node[TERM_IDS_KEY] = ids
        ids.append(entry.term_id)
    return root


def match_terms(tokens: Sequence[str], trie: Mapping[str, object]) -> Counter:
    found: Counter = Counter()
    token_count = len(tokens)
    for start_idx in range(token_count):
        first = tokens[start_idx]
        node = trie.get(first)
        if not isinstance(node, dict):
            continue

        ids = node.get(TERM_IDS_KEY)
        if isinstance(ids, list):
            for term_id in ids:
                found[term_id] += 1

        idx = start_idx + 1
        while idx < token_count:
            next_node = node.get(tokens[idx])
            if not isinstance(next_node, dict):
                break
            node = next_node
            ids = node.get(TERM_IDS_KEY)
            if isinstance(ids, list):
                for term_id in ids:
                    found[term_id] += 1
            idx += 1
    return found


def scan_datasets(dataset_paths: Sequence[Path], entries: Sequence[TermEntry]) -> Tuple[int, Dict[int, TermStats], Dict[str, AllergenSummary]]:
    trie = build_term_trie(entries)
    stats_by_term_id: Dict[int, TermStats] = {entry.term_id: TermStats() for entry in entries}
    allergen_summary: Dict[str, AllergenSummary] = {key: AllergenSummary() for key in ALLERGEN_ORDER}

    total_rows = 0

    for dataset_path in dataset_paths:
        with dataset_path.open("r", encoding="utf-8") as handle:
            for raw_line in handle:
                line = raw_line.strip()
                if not line:
                    continue
                payload = json.loads(line)
                if not isinstance(payload, dict):
                    continue

                total_rows += 1
                text = str(payload.get("text") or "")
                labels = {
                    allergen_key_from_label(str(item))
                    for item in (payload.get("allergens") or [])
                    if allergen_key_from_label(str(item))
                }

                for allergen_key in labels:
                    if allergen_key in allergen_summary:
                        allergen_summary[allergen_key].labeled_rows += 1

                tokens = tokenize_text(text)
                if not tokens:
                    continue

                matched_term_counts = match_terms(tokens, trie)
                if not matched_term_counts:
                    continue

                matched_by_allergen: Set[str] = set()

                for term_id, mentions in matched_term_counts.items():
                    entry = entries[term_id]
                    stat = stats_by_term_id[term_id]
                    stat.rows_matched += 1
                    stat.mentions_total += int(mentions)
                    if entry.allergen_key in labels:
                        stat.rows_with_target_allergen += 1
                    matched_by_allergen.add(entry.allergen_key)

                    summary = allergen_summary[entry.allergen_key]
                    summary.term_mentions += int(mentions)
                    summary.unique_term_ids.add(term_id)

                for allergen_key in labels:
                    if allergen_key in matched_by_allergen:
                        allergen_summary[allergen_key].labeled_rows_with_term += 1

    return total_rows, stats_by_term_id, allergen_summary


def term_sort_key(entry: TermEntry, stat: TermStats) -> Tuple[int, int, int, str]:
    allergen_idx = ALLERGEN_ORDER.index(entry.allergen_key) if entry.allergen_key in ALLERGEN_ORDER else 999
    return (allergen_idx, -stat.rows_with_target_allergen, -stat.rows_matched, entry.term_display.lower())


def write_term_frequency_csv(
    output_path: Path,
    entries: Sequence[TermEntry],
    stats_by_term_id: Mapping[int, TermStats],
    total_rows: int,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sorted_rows = sorted(entries, key=lambda entry: term_sort_key(entry, stats_by_term_id[entry.term_id]))
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "allergen",
                "ingredient_name",
                "source_type",
                "matched_rows",
                "target_allergen_rows",
                "non_target_rows",
                "target_precision_pct",
                "mentions_total",
                "prevalence_pct",
                "source_origin",
            ]
        )
        for entry in sorted_rows:
            stat = stats_by_term_id[entry.term_id]
            if stat.rows_matched <= 0:
                continue
            precision = (100.0 * stat.rows_with_target_allergen / stat.rows_matched) if stat.rows_matched else 0.0
            prevalence = (100.0 * stat.rows_matched / total_rows) if total_rows else 0.0
            writer.writerow(
                [
                    entry.allergen_name,
                    entry.term_display,
                    entry.source_type,
                    stat.rows_matched,
                    stat.rows_with_target_allergen,
                    stat.rows_matched - stat.rows_with_target_allergen,
                    f"{precision:.2f}",
                    stat.mentions_total,
                    f"{prevalence:.4f}",
                    entry.source_origin,
                ]
            )


def write_allergen_summary_csv(output_path: Path, summary: Mapping[str, AllergenSummary]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "allergen",
                "labeled_rows",
                "labeled_rows_with_any_term",
                "label_coverage_pct",
                "term_mentions_total",
                "unique_terms_matched",
            ]
        )
        for allergen_key in ALLERGEN_ORDER:
            item = summary.get(allergen_key) or AllergenSummary()
            coverage = (100.0 * item.labeled_rows_with_term / item.labeled_rows) if item.labeled_rows else 0.0
            writer.writerow(
                [
                    WORKBOOK_ALLERGEN_BY_KEY.get(allergen_key, allergen_key),
                    item.labeled_rows,
                    item.labeled_rows_with_term,
                    f"{coverage:.2f}",
                    item.term_mentions,
                    len(item.unique_term_ids or set()),
                ]
            )


def is_valid_added_expansion_term(term: TermEntry, stat: TermStats) -> bool:
    if term.source_type == "existing":
        return False
    if stat.rows_matched <= 0:
        return False
    if len(term.term_tokens) == 1 and term.term_tokens[0] in GENERIC_SINGLE_WORD_EXPANSION_BLOCKLIST:
        return False
    return True


def select_new_workbook_terms(
    entries: Sequence[TermEntry],
    stats_by_term_id: Mapping[int, TermStats],
    existing_token_keys: Set[Tuple[str, Tuple[str, ...]]],
    min_target_rows: int,
    min_precision: float,
    include_parenthetical_additions: bool,
) -> List[Tuple[TermEntry, TermStats]]:
    selected: List[Tuple[TermEntry, TermStats]] = []
    seen_selected: Set[Tuple[str, Tuple[str, ...]]] = set()

    for entry in entries:
        stat = stats_by_term_id[entry.term_id]
        if not is_valid_added_expansion_term(entry, stat):
            continue
        if (not include_parenthetical_additions) and entry.source_type == "expanded_parenthetical":
            continue
        token_key = (entry.allergen_key, entry.term_tokens)
        if token_key in existing_token_keys or token_key in seen_selected:
            continue
        if stat.rows_with_target_allergen < int(min_target_rows):
            continue
        precision = stat.rows_with_target_allergen / stat.rows_matched if stat.rows_matched else 0.0
        if precision < float(min_precision):
            continue
        selected.append((entry, stat))
        seen_selected.add(token_key)

    selected.sort(key=lambda item: term_sort_key(item[0], item[1]))
    return selected


def write_new_terms_csv(output_path: Path, selected: Sequence[Tuple[TermEntry, TermStats]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "allergen",
                "ingredient_name",
                "source_type",
                "target_allergen_rows",
                "matched_rows",
                "target_precision_pct",
                "mentions_total",
            ]
        )
        for entry, stat in selected:
            precision = (100.0 * stat.rows_with_target_allergen / stat.rows_matched) if stat.rows_matched else 0.0
            writer.writerow(
                [
                    entry.allergen_name,
                    entry.term_display,
                    entry.source_type,
                    stat.rows_with_target_allergen,
                    stat.rows_matched,
                    f"{precision:.2f}",
                    stat.mentions_total,
                ]
            )


def rebuild_sheet(workbook: openpyxl.Workbook, sheet_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
    if sheet_name in workbook.sheetnames:
        old = workbook[sheet_name]
        workbook.remove(old)
    return workbook.create_sheet(sheet_name)


def populate_frequency_sheet(
    workbook: openpyxl.Workbook,
    entries: Sequence[TermEntry],
    stats_by_term_id: Mapping[int, TermStats],
    total_rows: int,
) -> None:
    ws = rebuild_sheet(workbook, "Ingredient Frequency")
    ws.append(
        [
            "Allergen",
            "Ingredient Name",
            "Source Type",
            "Matched Label Rows",
            "Target-Allergen Rows",
            "Target Precision %",
            "Mentions Total",
            "Prevalence % (All Rows)",
            "Source Origin",
        ]
    )

    for entry in sorted(entries, key=lambda item: term_sort_key(item, stats_by_term_id[item.term_id])):
        stat = stats_by_term_id[entry.term_id]
        if stat.rows_matched <= 0:
            continue
        precision = (100.0 * stat.rows_with_target_allergen / stat.rows_matched) if stat.rows_matched else 0.0
        prevalence = (100.0 * stat.rows_matched / total_rows) if total_rows else 0.0
        ws.append(
            [
                entry.allergen_name,
                entry.term_display,
                entry.source_type,
                stat.rows_matched,
                stat.rows_with_target_allergen,
                round(precision, 2),
                stat.mentions_total,
                round(prevalence, 4),
                entry.source_origin,
            ]
        )
    ws.freeze_panes = "A2"


def populate_allergen_summary_sheet(workbook: openpyxl.Workbook, summary: Mapping[str, AllergenSummary]) -> None:
    ws = rebuild_sheet(workbook, "Allergen Frequency Summary")
    ws.append(
        [
            "Allergen",
            "Labeled Rows",
            "Rows With >=1 Ingredient Term",
            "Coverage %",
            "Term Mentions Total",
            "Unique Terms Matched",
        ]
    )

    for allergen_key in ALLERGEN_ORDER:
        item = summary.get(allergen_key) or AllergenSummary()
        coverage = (100.0 * item.labeled_rows_with_term / item.labeled_rows) if item.labeled_rows else 0.0
        ws.append(
            [
                WORKBOOK_ALLERGEN_BY_KEY.get(allergen_key, allergen_key),
                item.labeled_rows,
                item.labeled_rows_with_term,
                round(coverage, 2),
                item.term_mentions,
                len(item.unique_term_ids or set()),
            ]
        )
    ws.freeze_panes = "A2"


def append_new_terms_to_map_sheet(
    workbook: openpyxl.Workbook,
    selected_terms: Sequence[Tuple[TermEntry, TermStats]],
) -> None:
    ws_map = workbook["Allergen-Ingredient Map"]
    for entry, stat in selected_terms:
        precision = (100.0 * stat.rows_with_target_allergen / stat.rows_matched) if stat.rows_matched else 0.0
        ws_map.append(
            [
                entry.allergen_name,
                entry.term_display,
                "Auto-expanded term",
                f"Auto-added from dataset scan; target rows={stat.rows_with_target_allergen}, precision={precision:.1f}%",
                "Auto-generated from existing workbook terms + processed ingredient-label datasets",
            ]
        )


def resolve_dataset_paths(args_dataset: Sequence[str], repo_root: Path) -> List[Path]:
    candidate_paths = list(args_dataset) if args_dataset else list(DEFAULT_DATASET_FILES)
    resolved: List[Path] = []
    for raw in candidate_paths:
        path = Path(raw)
        if not path.is_absolute():
            path = repo_root / path
        resolved.append(path)
    missing = [path for path in resolved if not path.exists()]
    if missing:
        lines = "\n".join(f"- {path}" for path in missing)
        raise FileNotFoundError(f"Missing dataset file(s):\n{lines}")
    return resolved


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[2]

    workbook_input = Path(args.workbook_input)
    if not workbook_input.exists():
        print(f"Workbook not found: {workbook_input}")
        return 1

    dataset_paths = resolve_dataset_paths(args.dataset, repo_root)
    workbook, entries, existing_token_keys = build_term_inventory(workbook_input)

    total_rows, stats_by_term_id, allergen_summary = scan_datasets(dataset_paths, entries)

    selected_terms = select_new_workbook_terms(
        entries=entries,
        stats_by_term_id=stats_by_term_id,
        existing_token_keys=existing_token_keys,
        min_target_rows=int(args.min_added_target_rows),
        min_precision=float(args.min_added_precision),
        include_parenthetical_additions=bool(args.include_parenthetical_additions),
    )

    csv_output_dir = Path(args.csv_output_dir)
    if not csv_output_dir.is_absolute():
        csv_output_dir = repo_root / csv_output_dir
    csv_output_dir.mkdir(parents=True, exist_ok=True)

    term_csv = csv_output_dir / "allergen_ingredient_term_frequency.csv"
    summary_csv = csv_output_dir / "allergen_frequency_summary.csv"
    new_terms_csv = csv_output_dir / "allergen_new_terms_added.csv"

    write_term_frequency_csv(term_csv, entries, stats_by_term_id, total_rows)
    write_allergen_summary_csv(summary_csv, allergen_summary)
    write_new_terms_csv(new_terms_csv, selected_terms)

    append_new_terms_to_map_sheet(workbook, selected_terms)
    populate_frequency_sheet(workbook, entries, stats_by_term_id, total_rows)
    populate_allergen_summary_sheet(workbook, allergen_summary)

    workbook_output = Path(args.workbook_output)
    workbook_output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_output)

    existing_count = sum(1 for entry in entries if entry.source_type == "existing")
    expanded_count = len(entries) - existing_count
    matched_term_count = sum(1 for entry in entries if stats_by_term_id[entry.term_id].rows_matched > 0)

    print(f"Workbook input: {workbook_input}")
    print(f"Workbook output: {workbook_output}")
    print(f"Dataset rows scanned: {total_rows}")
    print(f"Existing terms: {existing_count}")
    print(f"Expanded candidate terms: {expanded_count}")
    print(f"Terms matched in dataset: {matched_term_count}")
    print(f"New workbook rows added: {len(selected_terms)}")
    print(f"Term frequency CSV: {term_csv}")
    print(f"Allergen summary CSV: {summary_csv}")
    print(f"New terms CSV: {new_terms_csv}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
